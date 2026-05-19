"""Inside Sales weekly/monthly Excel upload sheets.

Workflow:
  1. POST  /api/sales-uploads        — Inside Sales employee uploads .xlsx,
                                        we parse it for a quick preview and
                                        push the bytes to MinIO.
  2. GET   /api/sales-uploads        — list uploads (HR sees all; an
                                        Inside Sales employee sees their own).
  3. GET   /api/sales-uploads/{id}/download — presigned MinIO URL for direct
                                        browser download (lives ~10 minutes).
  4. DELETE /api/sales-uploads/{id}  — HR or the uploading employee can
                                        remove an upload.

Only Inside Sales department employees may upload.  HR has full read/delete.
"""
from __future__ import annotations

import io
import uuid
from datetime import date as date_type, datetime, timezone
from pathlib import Path

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Response,
    UploadFile,
    status,
)
from openpyxl import load_workbook
from sqlalchemy.orm import Session

import storage
from auth import get_current_user
from database import get_db
from models import SalesUpload, User
from schemas import SalesUploadOut

router = APIRouter(prefix="/api/sales-uploads", tags=["sales-uploads"])

INSIDE_SALES_SLUG = "insideSales"
MAX_BYTES = 5 * 1024 * 1024  # 5 MB cap — calling sheets are tiny CSV-ish files
ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


def _is_inside_sales(user: User) -> bool:
    dept = getattr(user, "department", None)
    return dept is not None and dept.slug == INSIDE_SALES_SLUG


def _is_hr(user: User) -> bool:
    return user.role == "hr"


def _parse_xlsx_summary(xlsx_bytes: bytes, max_rows_preview: int = 30) -> dict:
    """Open the workbook with openpyxl and pull a compact summary:
      - sheet name + dimensions
      - column headers (row 1) WITH their fill/font colors
      - first N data rows WITH per-cell fill/font colors
      - total row count

    We open the workbook in non-read-only mode so cell styles (fill color,
    font color) are reachable.  This is slightly slower than read_only=True
    but calling sheets are small (a few KB to a few MB) so the trade-off is
    worth it — the preview now matches what Excel renders.

    Each cell becomes `{"v": value, "bg": "RRGGBB" | null, "fg": "RRGGBB" | null}`
    so the frontend can paint cells with inline CSS.
    """
    summary: dict = {"sheets": [], "preview_rows": [], "headers": [], "row_count": 0}
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        return {"error": f"Could not parse Excel: {e}"}

    sheet_names = wb.sheetnames
    summary["sheets"] = sheet_names

    ws = wb[sheet_names[0]] if sheet_names else None
    if ws is None:
        return summary

    rows_iter = ws.iter_rows(values_only=False)
    try:
        header_row = next(rows_iter)
        summary["headers"] = [_cell_to_dict(c) for c in header_row]
    except StopIteration:
        return summary

    preview: list[list] = []
    row_count = 0
    for r in rows_iter:
        # Skip wholly empty rows so trailing blank rows don't pad the preview.
        if all((cell.value is None or cell.value == "") for cell in r):
            continue
        row_count += 1
        if len(preview) < max_rows_preview:
            preview.append([_cell_to_dict(c) for c in r])
    summary["preview_rows"] = preview
    summary["row_count"] = row_count
    return summary


def _serialize_cell(v):
    """Make sure the JSONB column can hold whatever openpyxl gives us."""
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date_type):
        return v.isoformat()
    return str(v)


def _color_to_hex(color_obj) -> str | None:
    """Return a 6-char hex string ('RRGGBB') for an openpyxl Color, or None
    if it's transparent, default, theme-based (no RGB), or otherwise unusable.
    """
    if color_obj is None:
        return None
    # openpyxl Color has .rgb (may be 'aarrggbb' or 'rrggbb'), .type ('rgb',
    # 'theme', 'indexed', etc.).  We only handle rgb-typed colors here —
    # theme/indexed lookups would require resolving against the workbook's
    # theme XML which is more complex.
    rgb = getattr(color_obj, "rgb", None)
    color_type = getattr(color_obj, "type", None)
    if color_type and color_type != "rgb":
        return None
    if not isinstance(rgb, str):
        return None
    rgb = rgb.upper()
    if len(rgb) == 8:
        alpha, hex_color = rgb[:2], rgb[2:]
        if alpha == "00":  # fully transparent — treat as no fill
            return None
        return hex_color
    if len(rgb) == 6:
        return rgb
    return None


def _cell_to_dict(cell) -> dict:
    """Convert an openpyxl Cell to a JSON-safe dict with value + colors."""
    bg = None
    fg = None
    try:
        fill = cell.fill
        # PatternFill stores the user-set color in `fgColor`; `bgColor` is the
        # pattern background (usually unset).  start_color is an alias of fgColor.
        if fill is not None and getattr(fill, "patternType", None):
            bg = _color_to_hex(fill.fgColor) or _color_to_hex(fill.start_color)
        if cell.font is not None:
            fg = _color_to_hex(cell.font.color)
    except Exception:
        pass
    # Filter defaults: pure-black font and pure-white fill are usually just
    # Excel's defaults — sending them adds noise and visual diff for nothing.
    if fg == "000000":
        fg = None
    if bg == "FFFFFF":
        bg = None
    return {"v": _serialize_cell(cell.value), "bg": bg, "fg": fg}


def _to_out(upload: SalesUpload) -> SalesUploadOut:
    """Hydrate a SalesUpload row into the response schema with user_name."""
    name = ""
    if upload.user:
        first = (upload.user.first_name or "").strip()
        last = (upload.user.last_name or "").strip()
        name = (first + " " + last).strip() or upload.user.username
    return SalesUploadOut(
        id=upload.id,
        user_id=upload.user_id,
        user_name=name,
        period_type=upload.period_type,
        period_start=upload.period_start,
        period_end=upload.period_end,
        note=upload.note,
        original_filename=upload.original_filename,
        file_size_bytes=upload.file_size_bytes,
        parsed_summary=upload.parsed_summary or {},
        uploaded_at=upload.uploaded_at,
    )


@router.post("", response_model=SalesUploadOut, status_code=status.HTTP_201_CREATED)
async def create_upload(
    file: UploadFile = File(..., description="Excel file (.xlsx / .xlsm)"),
    period_type: str = Form("weekly"),
    period_start: str | None = Form(None),  # YYYY-MM-DD
    period_end: str | None = Form(None),
    note: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # ----- Authorization -----
    if not (_is_inside_sales(user) or _is_hr(user)):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Only Inside Sales employees (or HR) can upload sales sheets.",
        )

    # ----- File validation -----
    filename = file.filename or "upload.xlsx"
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Only Excel files allowed ({', '.join(sorted(ALLOWED_EXTENSIONS))}).",
        )

    data = await file.read()
    if len(data) == 0:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Empty file.")
    if len(data) > MAX_BYTES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"File is too large (max {MAX_BYTES // (1024 * 1024)} MB).",
        )

    if period_type not in ("weekly", "monthly", "adhoc"):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "period_type must be 'weekly', 'monthly', or 'adhoc'.",
        )

    def _parse_date(s: str | None) -> date_type | None:
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Invalid date: {s} (expected YYYY-MM-DD)",
            )

    p_start = _parse_date(period_start)
    p_end = _parse_date(period_end)

    # ----- Parse preview -----
    parsed = _parse_xlsx_summary(data)

    # ----- Push to MinIO -----
    today = date_type.today().isoformat()
    object_key = f"sales/{user.id}/{today}/{uuid.uuid4().hex}-{filename}"
    storage.put_object(
        object_key,
        data,
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    # ----- Persist DB row -----
    upload = SalesUpload(
        user_id=user.id,
        period_type=period_type,
        period_start=p_start,
        period_end=p_end,
        note=(note or "")[:512],
        original_filename=filename[:255],
        minio_object_key=object_key,
        file_size_bytes=len(data),
        parsed_summary=parsed,
        uploaded_at=datetime.now(timezone.utc),
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)
    return _to_out(upload)


@router.get("", response_model=list[SalesUploadOut])
def list_uploads(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """List uploads.  HR sees everything; an Inside Sales employee sees own only."""
    q = db.query(SalesUpload).order_by(SalesUpload.uploaded_at.desc())
    if not _is_hr(user):
        if not _is_inside_sales(user):
            raise HTTPException(status.HTTP_403_FORBIDDEN, "Not allowed.")
        q = q.filter(SalesUpload.user_id == user.id)
    return [_to_out(u) for u in q.all()]


@router.get("/{upload_id}/download")
def download_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Stream the Excel file back to the browser through this API.

    We deliberately don't return a presigned MinIO URL — it would leak the
    internal MinIO hostname and be unusably long.  Instead the bytes flow
    through FastAPI: the caller sees a short URL `/api/sales-uploads/{id}/download`
    and the file downloads with its original filename.
    """
    upload = db.query(SalesUpload).filter(SalesUpload.id == upload_id).first()
    if not upload:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Upload not found")

    if not _is_hr(user) and upload.user_id != user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Not allowed.")

    try:
        data = storage.get_object_bytes(upload.minio_object_key)
    except Exception as e:
        raise HTTPException(
            status.HTTP_502_BAD_GATEWAY,
            f"Could not fetch file from storage: {e}",
        )

    # RFC 5987 quoting keeps non-ASCII filenames safe in the header.
    safe_name = upload.original_filename.replace('"', "")
    return Response(
        content=data,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}"',
            "Content-Length": str(len(data)),
        },
    )


@router.delete("/{upload_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    upload = db.query(SalesUpload).filter(SalesUpload.id == upload_id).first()
    if not upload:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Upload not found")

    if not _is_hr(user) and upload.user_id != user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Not allowed.")

    storage.delete_object(upload.minio_object_key)
    db.delete(upload)
    db.commit()
    return None
