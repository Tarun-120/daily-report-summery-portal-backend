"""
Import the real HR roster from an Excel sheet — SQLAlchemy version (no Django).

Run from inside the FastAPI container:
    docker compose exec fastapi python -m scripts.import_employees

By default DRY_RUN = True — first run prints what *would* happen without
touching the database.  Once you're happy, set DRY_RUN = False below and
re-run.

Notes:
- Custom department field templates (Sales / Marketing / Procurement /
  Logistics etc.) are NOT clobbered on re-runs.  Only newly-created
  departments get the generic 5-field template.
- Existing users are matched by (first_name, last_name) and updated in-place
  so re-running doesn't duplicate them or reset their password.
- Email pattern: firstname@acme.com, with a numeric suffix on collision.
- Password pattern: firstname@acme (PBKDF2-hashed before save).
"""
from __future__ import annotations

import re
import sys
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path

# Make `from auth import ...` etc. resolvable when run as a module.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from auth import hash_password  # noqa: E402
from database import SessionLocal  # noqa: E402
from models import DailyReport, Department, User  # noqa: E402
import model_events  # noqa: F401, E402  -- registers User invariants

# ============================================================================
# CONFIG — flip this to False when you're ready to actually write
# ============================================================================
DRY_RUN = True
EXCEL_PATH = "/app/employees.xlsx"

GENERIC_FIELDS = [
    {"key": "workDone", "label": "Work Done"},
    {"key": "workInProgress", "label": "Work in Progress"},
    {"key": "upcomingPriorities", "label": "Upcoming Priorities"},
    {"key": "challenges", "label": "Challenges Faced / Support Needed"},
    {"key": "otherUpdate", "label": "Other Update"},
]

# Excel-name -> (slug, display_name, color).  R&D variants with whitespace
# typos are unified into one slug; R&D-BESS is kept separate per the brief.
DEPT_MAP = {
    "BESS-Sales":  ("bessSales",  "BESS Sales",  "amber"),
    "Design":      ("design",     "Design",      "indigo"),
    "Finance":     ("finance",    "Finance",     "emerald"),
    "HR":          ("hrDept",     "HR",          "rose"),
    "Logistics":   ("logistics",  "Logistics",   "sky"),
    "Marketing":   ("marketing",  "Marketing",   "indigo"),
    "O & M":       ("om",         "O & M",       "zinc"),
    "Procurement": ("procurement", "Procurement", "emerald"),
    "Production":  ("production", "Production",  "amber"),
    "Project":     ("project",    "Project",     "rose"),
    "R & D":       ("rd",         "R & D",       "indigo"),
    "R& D":        ("rd",         "R & D",       "indigo"),  # typo alias
    "R & D-BESS":  ("rdBess",     "R & D-BESS",  "sky"),
    "Sales":       ("sales",      "Sales",       "rose"),
    "Service":     ("service",    "Service",     "emerald"),
    "Support":     ("support",    "Support",     "amber"),
    "Web Dev":     ("webDev",     "Web Dev",     "indigo"),
}


# ---------------------------------------------------------------------------

def _slugify_local(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _split_name(full: str) -> tuple[str, str]:
    parts = (full or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _to_date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    return v


def _read_excel():
    import openpyxl  # lazy import so the script gives a clear error if missing
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    rows = []

    # Sheet 1: Company A  cols: Organisation, S.No, Name, Dept, RM, DOJ
    ws1 = wb["Company A"]
    for row in ws1.iter_rows(min_row=2, values_only=True):
        org, _sn, name, dept, rm, doj = row
        if not name:
            continue
        rows.append({
            "name": str(name).strip(),
            "dept_excel": (dept or "").strip(),
            "title": "",
            "organisation": (org or "").strip(),
            "reporting_manager": (rm or "").strip(),
            "date_of_joining": _to_date(doj),
        })

    # Sheet 2: Company B  cols: NAME, EMP CODE, DESIGNATION, DEPT, RM, DOJ
    ws2 = wb["Company B"]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        name, _code, des, dept, rm, doj = row
        if not name:
            continue
        rows.append({
            "name": str(name).strip(),
            "dept_excel": (dept or "").strip(),
            "title": (des or "").strip(),
            "organisation": "Company B",
            "reporting_manager": (rm or "").strip(),
            "date_of_joining": _to_date(doj),
        })
    return rows


def _step(label: str) -> None:
    bar = "=" * 70
    print()
    print(bar)
    print(label)
    print(bar)


def main() -> None:
    print(f"\nDRY_RUN = {DRY_RUN}    (set DRY_RUN = False at top of script to write)")

    excel_rows = _read_excel()
    _step("1. Reading Excel")
    print(f"  {len(excel_rows)} rows with a name")

    unknown = sorted({r["dept_excel"] for r in excel_rows if r["dept_excel"] not in DEPT_MAP})
    if unknown:
        print("  ERROR: Unknown department names in Excel (add to DEPT_MAP):")
        for d in unknown:
            print(f"    - {d!r}")
        return

    db = SessionLocal()
    try:
        # ---- Plan ----
        _step("2. Plan: ensure departments exist")
        wanted = OrderedDict()
        for excel_name, (slug, display, color) in DEPT_MAP.items():
            wanted[slug] = (display, color)
        existing_slugs = {d.slug for d in db.query(Department).all()}
        for slug, (display, color) in wanted.items():
            print(f"  {'exists' if slug in existing_slugs else 'WILL CREATE'} {slug:<14} -> {display}")

        _step("3. Plan: import employees")
        # Match existing users by full name (case-insensitive) so re-runs update.
        existing_by_name: dict[tuple[str, str], User] = {}
        for u in db.query(User).filter(User.role == "employee").all():
            key = (u.first_name.strip().lower(), u.last_name.strip().lower())
            existing_by_name[key] = u

        taken_emails = {
            (e or "").lower()
            for e in db.query(User.email).filter(User.role.in_(["hr", "employee"])).all()
            for e in [e[0]] if e
        }

        plan = []
        for r in excel_rows:
            first, last = _split_name(r["name"])
            local = _slugify_local(first)
            if not local:
                print(f"  SKIP (no usable first name): {r['name']!r}")
                continue
            existing = existing_by_name.get((first.lower(), last.lower()))
            if existing:
                email = existing.email
            else:
                cand = f"{local}@acme.com"
                n = 2
                while cand in taken_emails:
                    cand = f"{local}{n}@acme.com"
                    n += 1
                taken_emails.add(cand)
                email = cand
            plan.append({
                "first": first, "last": last, "email": email,
                "password": f"{local}@acme",
                "dept_slug": DEPT_MAP[r["dept_excel"]][0],
                "title": r["title"],
                "organisation": r["organisation"],
                "reporting_manager": r["reporting_manager"],
                "date_of_joining": r["date_of_joining"],
                "existing": existing,
            })

        creates = sum(1 for p in plan if not p["existing"])
        updates = sum(1 for p in plan if p["existing"])
        print(f"  Plan: {creates} new, {updates} update existing  (total {len(plan)})")
        for p in plan[:5]:
            tag = "UPD" if p["existing"] else "NEW"
            doj = p["date_of_joining"].isoformat() if p["date_of_joining"] else "—"
            print(f"    [{tag}] {p['first']} {p['last']:<22} {p['email']:<32} dept={p['dept_slug']:<11} doj={doj}")
        if len(plan) > 5:
            print(f"    ... +{len(plan) - 5} more")

        if DRY_RUN:
            _step("DRY RUN COMPLETE — nothing was written")
            print("Set DRY_RUN = False at top of script to apply.")
            return

        # ---- Execute ----
        _step("4. Executing")

        # Departments — create if missing, never overwrite report_fields on update.
        slug_to_dept: dict[str, Department] = {}
        for slug, (display, color) in wanted.items():
            dept = db.query(Department).filter(Department.slug == slug).first()
            if dept is None:
                dept = Department(
                    slug=slug, name=display, color=color,
                    report_fields=GENERIC_FIELDS,
                )
                db.add(dept)
                db.flush()
                print(f"  created dept: {slug}")
            else:
                # Only sync display fields, leave report_fields alone.
                dept.name = display
                dept.color = color
                print(f"  kept    dept: {slug}")
            slug_to_dept[slug] = dept

        # Employees — create or update.
        now = datetime.now(timezone.utc)
        created_count = 0
        updated_count = 0
        for p in plan:
            user = p["existing"]
            is_new = user is None
            if is_new:
                user = User(
                    username=p["email"].split("@")[0],
                    email=p["email"],
                    is_active=True,
                    is_staff=False,
                    is_superuser=False,
                    role="employee",
                    date_joined=now,
                    password=hash_password(p["password"]),
                )
                db.add(user)
            user.first_name = p["first"]
            user.last_name = p["last"]
            user.title = p["title"] or ""
            user.department = slug_to_dept[p["dept_slug"]]
            user.organisation = p["organisation"] or ""
            user.reporting_manager = p["reporting_manager"] or ""
            user.date_of_joining = p["date_of_joining"]
            if is_new:
                created_count += 1
            else:
                updated_count += 1

        db.commit()
        print(f"  Employees created: {created_count}, updated: {updated_count}")

        _step("DONE")
        print(f"  Total employees: {db.query(User).filter(User.role == 'employee').count()}")
        print(f"  Total departments: {db.query(Department).count()}")
        print(f"  Total reports: {db.query(DailyReport).count()}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
